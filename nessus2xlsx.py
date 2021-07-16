#!/usr/bin/python
# encoding: utf-8
# Transform Nessus xml to Excel
# Authors: Benjamin Kellermann, Björn Kirschner, Jan Rude
# License: GPLv3
import sys
import os
import re
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.filters import SortCondition, SortState, AutoFilter
from openpyxl import load_workbook

if not os.path.exists(sys.argv[1]):
	sys.exit('ERROR: Input %s was not found!' % sys.argv[1])

# output file
file_name = os.path.basename(sys.argv[1])
output_file = file_name.replace(".nessus", ".xlsx")

# guess title
sheet_title = os.path.basename(sys.argv[1])
sheet_title = re.sub(r"\.nessus$", "", sheet_title)         # remove fileending
sheet_title = re.sub(r"_[0-9a-zA-Z]{6}$", "", sheet_title)  # remove random-string added by nessus
sheet_title = re.sub(r"_", " ", sheet_title)                # replace _ with " "

existing_tables = []
if os.path.exists(output_file):
	print('WARNING: Output %s already exists, adding a sheet.' % output_file)
	wb = load_workbook(output_file)
	for ws in wb.worksheets:
		for tbl in ws._tables:
			existing_tables.append(tbl.name)
	vulns_wb = wb.create_sheet()
else:
	# Open up a new excel file
	wb = Workbook()
	vulns_wb = wb.active


vulns_wb.title = sheet_title
vulns_wb.append(["Hostname", "Port", "Severity", "MGM Rating", "Comment", "Name", "Description", "Output", "Solution", "See Also"])

ns = {'cm': 'http://www.nessus.org/cm'}

severity_levels = {
	0: 'None',
	1: 'Low',
	2: 'Medium',
	3: 'High',
	4: 'Critical'
}

xml = ET.parse(sys.argv[1])
root = xml.getroot()

# Add a default style with striped rows and banded columns
style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)

# Check if report has compliance checks in it; ignore PCI DSS entries
compliance_wb = ''
for report_item in root.iter('ReportItem'):
	if "Compliance" in report_item.get('pluginName') and not "PCI DSS" in report_item.get('pluginName'):
		compliance_wb = wb.create_sheet()
		vulns_wb.title = ("%s (Vulnerabilities)" % sheet_title)
		compliance_wb.title = ("%s (Compliance Checks)" % sheet_title)
		compliance_wb.append(["Hostname", "Result", "Name", "Output", "Description", "Solution", "See Also"])
		break

# Iterate through findings
for host in root.iter('ReportHost'):
	hostname = host.get('name')

	for report_item in host.iter('ReportItem'):
		severity = severity_levels.get(int(report_item.get('severity')))
		description = report_item.find('description').text if not (report_item.find('description') is None) else ""

		pluginFamily = report_item.get('pluginFamily')
		# Compliance Check
		if pluginFamily == ("Policy Compliance" or "Settings"):
			pci = report_item.get('pluginName')
			if not "pci dss compliance" in pci.lower():
				name = report_item.find('cm:compliance-check-name', ns).text
				compliance_result = report_item.find('cm:compliance-result', ns)
				result = compliance_result.text
				actual_value = report_item.find('cm:compliance-actual-value', ns)
				if actual_value is None:  # usually when PASSED
					output = ''
				else:
					output = actual_value.text
				actual_solution = report_item.find('cm:compliance-solution', ns)
				if actual_solution is None:  # usually when PASSED
					solution = ''
				else:
					solution = actual_solution.text

				desc_pre = description.find('\n')
				desc_past = description.find('\nReference(s) : ')
				description_final = description[desc_pre+2:desc_past-1]

				temp_see_also = report_item.find('cm:compliance-see-also', ns)
				if temp_see_also is None:  # usually when PASSED
					see_also = ''
				else:
					see_also = temp_see_also.text

				compliance_wb.append([hostname, result, name, output, description_final, solution, see_also])
		else:
			port = (report_item.get('protocol') + "/" + report_item.get('port'))
			name = report_item.find('plugin_name').text
			if not "pci dss compliance" in name.lower(): # ignore 'PCI DSS Settings' entries
				output = report_item.find('plugin_output')
				if output is not None and output.text is not None:
					output = output.text.lstrip().rstrip()
				else:
					output = ""
				solution = report_item.find('solution').text if not (report_item.find('solution') is None) else ""
				see_also = report_item.find('see_also').text if not (report_item.find('see_also') is None) else ""
				see_also = see_also.encode()

				vulns_wb.append([hostname, port, severity, '', '', name, description, output, solution, see_also])

## Style options
# Host column
vulns_wb.column_dimensions['A'].width = 15
for host in vulns_wb['A']:
	host.alignment = Alignment(horizontal='center', vertical='top')

# Port column
vulns_wb.column_dimensions['B'].width = 10
for host in vulns_wb['B']:
	host.alignment = Alignment(horizontal='center', vertical='top')

# Severity column
vulns_wb.column_dimensions['C'].width = 14
for severity in vulns_wb['C']:
	severity.alignment = Alignment(horizontal='center', vertical='top')

# MGM rating column
vulns_wb.column_dimensions['D'].width = 18
for name in vulns_wb['D']:
	name.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

# Severity + MGM rating-styling
size=len(vulns_wb['D'])
vulns_wb.conditional_formatting.add('C2:D{}'.format(size), CellIsRule(operator='equal', fill=PatternFill(bgColor='ffff99', fill_type='solid'), formula=['"Low"']))
vulns_wb.conditional_formatting.add('C2:D{}'.format(size), CellIsRule(operator='equal', fill=PatternFill(bgColor='fabf8f', fill_type='solid'), formula=['"Medium"']))
vulns_wb.conditional_formatting.add('C2:D{}'.format(size), CellIsRule(operator='equal', fill=PatternFill(bgColor='d99594', fill_type='solid'), formula=['"High"']))
vulns_wb.conditional_formatting.add('C2:D{}'.format(size), CellIsRule(operator='equal', fill=PatternFill(bgColor='ff000000', fill_type='solid'), font=Font(color='ffffff'), formula=['"Critical"']))

# Comment
vulns_wb.column_dimensions['E'].width = 30
for comment in vulns_wb['E']:
	comment.alignment = Alignment(vertical='top', wrap_text=True)

# Name, Description, Output, Solution, See Also column
for column in ['F', 'G', 'H', 'I', 'J']:
	vulns_wb.column_dimensions[column].width = 50
	for col in vulns_wb[column]:
		col.alignment = Alignment(vertical='top', wrap_text=True)

# Center first row
for cell in vulns_wb['A':get_column_letter(vulns_wb.max_column)]:
	cell[0].alignment = Alignment(horizontal='center')

# equalize height
for i in range(vulns_wb.max_row):
	vulns_wb.row_dimensions[i+1].height = 15

# Get Table dimensions
if vulns_wb.max_row == 1:  # if table is empty
	dim = "A1:J2"
else:
	dim = vulns_wb.dimensions

# Add sorting filter
vulnsort = "Critical, High, Medium, Low, None"
sort = [
	SortCondition(ref='D:D', customList=vulnsort),  # MGM Rating
	SortCondition(ref='C:C', customList=vulnsort),  # Nessus rating
	SortCondition(ref='F:F'),                       # Plugin Name
	SortCondition(ref='H:H'),                       # Output -> useful because often similar things are reported
	SortCondition(ref='A:A'),                       # Hostname
	SortCondition(ref='B:B')                        # Port
]
ss = SortState(ref=dim, sortCondition=sort)
af = AutoFilter(ref=dim, sortState=ss)

tab_title = "%s_Vulnerabilities" % re.sub(r"[^a-zA-Z0-9]", "_", sheet_title)
while tab_title in existing_tables:
	tab_title += '1'
tab = Table(displayName=tab_title, ref=dim, sortState=ss, autoFilter=af)
tab.tableStyleInfo = style
vulns_wb.add_table(tab)

if compliance_wb:  # {{{
	# Hostname column
	compliance_wb.column_dimensions['A'].width = 15
	for host in compliance_wb['A']:
		host.alignment = Alignment(horizontal='center', vertical='top')

	# compliance check result colors
	compliance_wb.column_dimensions['B'].width = 12
	for severity in compliance_wb['B']:
		if severity.value == "PASSED":
			severity.fill = PatternFill(fgColor='32CD32', fill_type='solid')
		elif severity.value == "ERROR":
			severity.fill = PatternFill(fgColor='FFEE08', fill_type='solid')
		elif severity.value == "FAILED":
			severity.fill = PatternFill(fgColor='FF0000', fill_type='solid')
		severity.alignment = Alignment(horizontal='center', vertical='top')

	# Name, Output column
	for column in ['C', 'D']:
		compliance_wb.column_dimensions[column].width = 30
		for col in compliance_wb[column]:
			col.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

	# Description, Solution column
	for column in ['E', 'F']:
		compliance_wb.column_dimensions[column].width = 70
		for col in compliance_wb[column]:
			col.alignment = Alignment(vertical='top', wrap_text=True)

	# 'See Also' column
	compliance_wb.column_dimensions['G'].width = 40
	for see in compliance_wb['G']:
		see.alignment = Alignment(vertical='top', wrap_text=True)

	# Center first row
	for cell in compliance_wb['A':get_column_letter(compliance_wb.max_column)]:
		cell[0].alignment = Alignment(horizontal='center')

	# equalize height
	for i in range(compliance_wb.max_row):
		compliance_wb.row_dimensions[i+1].height = 15

	tab_title = "%s_Compliance" % re.sub(r"[^a-zA-Z0-9]", "_", sheet_title)
	while tab_title in existing_tables:
		tab_title += '1'
	tab = Table(displayName=tab_title, ref="A1:{}{}".format(get_column_letter(compliance_wb.max_column), compliance_wb.max_row))
	tab.tableStyleInfo = style
	compliance_wb.add_table(tab)
# }}}
# Save the file
wb.save(filename=output_file)
